# -*- coding: utf-8 -*-
# =================================================================================================
#   Quality Automation Script (Version 17.4 - Simplified Avg Point Trigger)
#   Description: This version simplifies the email filter logic. Instead of checking proximity
#   to the standard point, it now triggers an email for a PASS report if the average
#   point is above a simple, direct threshold (e.g., >=10), which is configurable
#   in master.json for full user control.
# =================================================================================================

import os
import sys
import json
import logging
import re
import shutil
from datetime import datetime
from pathlib import Path
from collections import defaultdict
from typing import List, Dict, Any, Optional, Tuple

# --- Required Library Imports & Error Handling ---
try:
    import pandas as pd
    import openpyxl
    import xlwings as xw
    from openpyxl.utils import column_index_from_string, range_boundaries
    import win32com.client as win32
    from dotenv import load_dotenv
except ImportError as e:
    print(f"\nFATAL: A required library is not found: '{e.name}'.")
    print("Please run this command in your terminal to install necessary libraries:")
    print("pip install pandas openpyxl xlwings pywin32 python-dotenv")
    sys.exit(1)


# ==========================================================
#                      Configuration & Setup
# ==========================================================

def setup_logging():
    """Sets up the logging system for both file and console output."""
    log_file = "automation_log.txt"
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, mode='w', encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    logging.info(f"Logging system initialized. Log file: {log_file}")


class ConfigLoader:
    """A class to load and access settings from a JSON file."""

    def __init__(self, file_path: str):
        self.config = self._load_config(file_path)

    def _load_config(self, file_path: str) -> Dict[str, Any]:
        """Loads configuration from a JSON file."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except FileNotFoundError:
            logging.critical(f"❌ FATAL: Configuration file '{file_path}' not found.")
            sys.exit(1)
        except json.JSONDecodeError:
            logging.critical(f"❌ FATAL: JSON syntax error in '{file_path}'.")
            sys.exit(1)

    def get(self, key: str, default: Any = None) -> Any:
        """Accesses nested configuration values using dot notation."""
        keys = key.split('.')
        value = self.config
        for k in keys:
            if isinstance(value, dict) and k in value:
                value = value[k]
            else:
                return default
        return value


# =============================================================================
#   Helper Functions
# =============================================================================

def get_excel_files(folder_path: Path) -> List[Path]:
    """Finds all .xlsx and .xlsm files in a specified folder."""
    if not folder_path.is_dir():
        logging.warning(f"Warning: The folder '{folder_path}' does not exist.")
        return []
    return list(folder_path.rglob("*.xlsx")) + list(folder_path.rglob("*.xlsm"))


def safe_float(value: Any) -> float:
    """Safely converts a value to a float, returning 0.0 if conversion fails."""
    if value is None:
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


# =============================================================================
#   TASK 1: Data Entry Handler
# =============================================================================

class DataEntryHandler:
    """Handles the data entry task into the main Excel workbook using xlwings to preserve shapes."""

    def __init__(self, config: ConfigLoader):
        self.config = config
        self.main_file_path = Path(config.get('paths.main_workbook'))
        self.backup_dir = Path(config.get('paths.main_workbook_backup'))
        self.mappings = config.get('mappings_data_entry')
        self.summary_mapping = self.mappings.get('summary_mapping', {})
        self.defect_mapping = self.mappings.get('defect_mapping', {})
        self.table_name = self.mappings.get('target_table_name', 'Table13')
        self.cell_map = config.get('cell_map_organization')

    def _clear_serial_numbers(self, ws):
        """Clears the serial number column of the target table before data entry."""
        try:
            table_ref = ws.api.ListObjects(self.table_name).Range.Address
            if not table_ref:
                logging.warning(f"Table '{self.table_name}' not found. Could not clear serial numbers.")
                return

            logging.info("Clearing serial number column ('C') of the table...")
            table_range = ws.range(table_ref)
            sl_no_column_range = table_range.columns[2].offset(row_offset=1).resize(row_size=table_range.rows.count - 1)
            sl_no_column_range.clear_contents()
            logging.info("Serial number column cleared successfully.")
        except Exception as e:
            logging.error(f"Error while clearing serial numbers: {e}")

    def _get_sorting_keys_from_file(self, file_path: Path) -> Tuple[str, int, str, int]:
        """Reads sorting keys from a file according to the rule: Buyer > Consignment > Result > Rolls."""
        wb = None
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = wb[self.cell_map['sheet_name']]
            buyer = str(sheet[self.cell_map['buyer']].value or '').strip()
            consignment_val = str(sheet[self.cell_map['consignment']].value or '').strip()
            consignment_num = 0
            if consignment_val:
                numeric_part = ''.join(re.findall(r'\d+', consignment_val))
                if numeric_part:
                    consignment_num = int(numeric_part)
            result = str(sheet[self.cell_map['result']].value or '').strip()
            rolls_val = sheet[self.cell_map['rolls']].value
            rolls = int(safe_float(rolls_val))
            return (buyer, consignment_num, result, rolls)
        except Exception as e:
            logging.error(f"Error reading sorting key from '{file_path.name}': {e}")
            return ('', 0, '', 0)
        finally:
            if wb:
                wb.close()

    def _extract_data_for_entry(self, source_path: Path) -> Optional[Dict[str, Any]]:
        """Extracts data from the source file for entry using openpyxl for speed."""
        workbook = None
        try:
            workbook = openpyxl.load_workbook(filename=source_path, read_only=True, data_only=True)
            entry_data = {}
            summary_sheet = workbook["Summary"]
            for source_cell, target_column in self.summary_mapping.items():
                entry_data[target_column] = summary_sheet[source_cell].value

            normalized_defect_mapping = {k.lower().strip(): v for k, v in self.defect_mapping.items()}
            defect_points = {col: 0 for col in self.defect_mapping.values()}
            unmatched_defect_points = 0

            visible_pages = [s for s in workbook.worksheets if
                             s.title.startswith("Page ") and s.sheet_state == 'visible']
            for sheet in visible_pages:
                for row_num in range(23, 39):
                    defect_name = str(sheet[f'A{row_num}'].value or '').strip()
                    if not defect_name: continue
                    current_row_sum = sum(safe_float(sheet.cell(row=row_num, column=col).value) for col in
                                          range(column_index_from_string('V'), column_index_from_string('AO') + 1))
                    target_col = normalized_defect_mapping.get(defect_name.lower())
                    if target_col:
                        defect_points[target_col] = defect_points.get(target_col, 0) + current_row_sum
                    else:
                        unmatched_defect_points += current_row_sum

            entry_data.update(defect_points)
            entry_data['AK'] = unmatched_defect_points
            return entry_data
        except Exception as e:
            logging.error(f"  - Error processing file {source_path.name}: {e}")
            return None
        finally:
            if workbook:
                workbook.close()

    def run(self, source_files_list: List[Path]) -> int:
        """Executes the data entry process using xlwings to preserve file integrity."""
        logging.info("=" * 50 + "\nTASK 1: Starting Data Entry into Main Workbook...\n" + "=" * 50)
        if not source_files_list:
            logging.warning("No files found for data entry.")
            return 0

        logging.info("Collecting and sorting data from files...")
        all_data_to_enter = []
        for file in source_files_list:
            logging.info(f"   -> Reading data from: {file.name}")
            data = self._extract_data_for_entry(file)
            if data:
                sort_keys = self._get_sorting_keys_from_file(file)
                all_data_to_enter.append({'data': data, 'sort_keys': sort_keys, 'file_name': file.name})

        sorted_data = sorted(all_data_to_enter, key=lambda x: x['sort_keys'])
        logging.info("Data sorted successfully.")

        if not self.main_file_path.is_file():
            logging.error(f"Main workbook not found at: '{self.main_file_path}'")
            return 0

        excel_app = None
        try:
            excel_app = xw.App(visible=False)
            target_workbook = excel_app.books.open(self.main_file_path)
            ws = target_workbook.sheets["Data Analysis report"]

            self._clear_serial_numbers(ws)

            table_ref = ws.api.ListObjects(self.table_name).Range.Address
            if not table_ref:
                logging.error(f"Target table '{self.table_name}' not found. Cannot perform data entry.")
                return 0

            table_range = ws.range(table_ref)
            invoice_col = table_range.columns[5]  # Column F
            first_empty_cell = None
            for cell in invoice_col.offset(row_offset=1).resize(row_size=table_range.rows.count - 1 + 500):
                if cell.value is None:
                    first_empty_cell = cell
                    break

            next_row = first_empty_cell.row if first_empty_cell else ws.cells.last_cell.row + 1
            if next_row < 181: next_row = 181

            logging.info(f"Data entry will start from row {next_row}.")
            sl_no_counter = 1
            for item in sorted_data:
                entry = item['data']
                logging.info(f"   -> Writing data for: {item['file_name']}")
                ws.range(f'C{next_row}').value = sl_no_counter
                for col_letter, value in entry.items():
                    try:
                        ws.range(f'{col_letter}{next_row}').value = value
                    except Exception as cell_error:
                        logging.warning(f"Could not write to row {next_row}, column {col_letter}: {cell_error}")
                next_row += 1
                sl_no_counter += 1

            target_workbook.save()
            logging.info("Data entry completed successfully!")

            try:
                self.backup_dir.mkdir(exist_ok=True)
                backup_path = self.backup_dir / self.main_file_path.name
                shutil.copy2(self.main_file_path, backup_path)
                logging.info(f"A backup of the main file was successfully created at '{backup_path}'.")
            except Exception as backup_error:
                logging.error(f"Error creating backup: {backup_error}")

            return len(sorted_data)
        except Exception as e:
            logging.error(f"An unexpected error occurred during data entry: {e}", exc_info=True)
            return 0
        finally:
            if excel_app:
                for book in excel_app.books:
                    book.close()
                excel_app.quit()


# =============================================================================
#   TASK 2: Emailer
# =============================================================================

class Emailer:
    """Handles email creation and drafting with advanced data analysis."""

    def __init__(self, config: ConfigLoader):
        self.config = config
        self.cell_map = config.get('cell_map_organization')
        self.email_settings = config.get('email_settings')
        self.triggers = config.get('email_filter_rules.pass_report_triggers', {})
        self.review_folder = Path(config.get('paths.manual_review'))
        self.primary_recipient = self.email_settings.get('primary_recipient')
        self.secondary_recipient = self.email_settings.get('secondary_recipient')

    def _is_critical_shading(self, val: Any) -> bool:
        """Helper to determine if a shade value is critical (<=4)."""
        if val is None: return False
        str_val = str(val).strip()
        if not str_val: return False

        try:
            if "/" in str_val:
                first_part = str_val.split('/')[0]
                return safe_float(first_part) < 4
            else:
                return safe_float(str_val) <= 4
        except (ValueError, IndexError):
            return False

    def _analyze_report_data(self, file_path: Path) -> Dict[str, Any]:
        """Performs a deep analysis of a report file to find potential issues."""
        analysis = {"send_reason": "N/A"}
        wb = None
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            summary = wb[self.cell_map['sheet_name']]

            # --- Safely Get Raw Values ---
            order_width = safe_float(summary[self.cell_map['order_width']].value)
            actual_width = safe_float(summary[self.cell_map['actual_width']].value)
            ticked_yards = safe_float(summary[self.cell_map['ticked_yards']].value)
            short_excess = safe_float(summary[self.cell_map['total_short_excess']].value)
            avg_point = safe_float(summary[self.cell_map['avg_point']].value)
            check_roll = int(safe_float(summary[self.cell_map['check_roll']].value))

            # --- Width Shortage Check ---
            width_diff = order_width - actual_width
            if actual_width > 0 and order_width > 0 and width_diff > self.triggers.get('width_shortage_tolerance_inch',
                                                                                       0.5):
                analysis[
                    "send_reason"] = f"Width shortage > {self.triggers.get('width_shortage_tolerance_inch', 0.5)}\""
                return analysis

            # --- Length Shortage Check ---
            if ticked_yards > 0 and short_excess < 0:
                length_percent = (abs(short_excess) / ticked_yards) * 100
                if length_percent >= self.triggers.get('length_shortage_percentage', 0.5):
                    analysis[
                        "send_reason"] = f"Length shortage >= {self.triggers.get('length_shortage_percentage', 0.5)}%"
                    return analysis

            # --- Direct Average Point Check ---
            if avg_point >= self.triggers.get('avg_point_threshold', 10):
                analysis[
                    "send_reason"] = f"Avg point {avg_point} >= threshold {self.triggers.get('avg_point_threshold', 10)}"
                return analysis

            # --- Shading Percentage Check (Only on visible sheets) ---
            if check_roll > 0:
                critical_shade_rolls = 0
                visible_pages = [s for s in wb.worksheets if s.title.startswith("Page ") and s.sheet_state == 'visible']
                for sheet in visible_pages:
                    last_col = sheet.max_column
                    for i in range(2, last_col + 1, 4):
                        is_roll_critical = False
                        for j in range(i, i + 4):
                            if j > last_col: break
                            for k in range(15, 18):
                                if self._is_critical_shading(sheet.cell(row=k, column=j).value):
                                    is_roll_critical = True
                                    break
                            if is_roll_critical: break
                        if is_roll_critical: critical_shade_rolls += 1

                shading_percent = (critical_shade_rolls / check_roll) * 100
                if shading_percent >= self.triggers.get('shading_percentage_threshold', 15):
                    analysis[
                        "send_reason"] = f"Critical shading >= {self.triggers.get('shading_percentage_threshold', 15)}%"
                    return analysis

            return analysis
        except Exception as e:
            logging.error(f"Could not analyze data for '{file_path.name}': {e}")
            return {"send_reason": "Analysis Error"}
        finally:
            if wb:
                wb.close()

    def _get_report_data(self, file_path: Path) -> Optional[Dict]:
        """Collects basic data for the email from the Excel file."""
        wb = None
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet_name = self.config.get('cell_map_organization.sheet_name')
            sheet = wb[sheet_name]
            data = {}
            for key, cell in self.config.get('cell_map_organization').items():
                if key != 'sheet_name':
                    data[key] = str(sheet[cell].value or '').strip()
            data['file_path'] = file_path
            return data
        except Exception as e:
            logging.error(f"Could not read data from '{file_path.name}': {e}")
            return None
        finally:
            if wb:
                wb.close()

    def _classify_report(self, report: Dict) -> str:
        """Classifies a report as 'SEND' or 'REVIEW' based on advanced analysis."""
        result_lower = report.get('result', '').lower()

        if 'fail' in result_lower or 'rejected' in result_lower:
            return 'SEND'

        if 'pass' in result_lower:
            analysis_result = self._analyze_report_data(report['file_path'])
            if analysis_result["send_reason"] != "N/A":
                logging.info(
                    f"   - '{report['file_path'].name}' is a PASS report, but will be sent. Reason: {analysis_result['send_reason']}.")
                return 'SEND'

        logging.info(f"   - '{report['file_path'].name}' is a standard PASS report. It will be skipped.")
        return 'REVIEW'

    def _generate_email_body(self, buyer: str, supplier: str, reports: List[Dict]) -> str:
        """Generates the HTML body for the email, grouped by style."""
        body = f"""
        <html><head><style>
            body {{ font-family: Calibri, sans-serif; font-size: 11pt; }}
            .fail-text {{ color: red; font-weight: bold; }}
            .pass-text {{ color: green; }}
            .container {{ margin-bottom: 8px; }}
            .icon {{ font-size: 1.1em; }}
        </style></head><body>
            <p>Dear Concern,</p>
            <p>Please find the attached Fabric Inspection Report(s). The summary is mentioned below:</p>
            <p style="margin: 0;"><b>Buyer:</b> {buyer}</p>
            <p style="margin: 10px 0;"><b>Supplier:</b> {supplier}</p><hr>
        """

        reports_by_style = defaultdict(list)
        for r in reports:
            reports_by_style[r.get('style', 'N/A')].append(r)

        for style, style_reports in reports_by_style.items():
            body += f'<p style="margin-top: 15px; margin-bottom: 5px;"><b>Style:</b> {style}</p>'
            body += '<div style="margin-top: 5px; padding-left: 25px;">'
            for report in style_reports:
                result = report.get('result', 'N/A')
                comment = report.get('comment', '')
                result_class = "fail-text" if any(w in result.lower() for w in ['fail', 'rejected']) else "pass-text"
                reason_text = f"Due to: {comment}" if comment else ""

                body += f"""
                <div class="container">
                    <span class="icon">➢</span> 
                    <b>{report.get('color', 'N/A')}</b> ({report.get('rolls', 'N/A')} Rolls): 
                    <span class="{result_class}">{result.upper()}</span> {reason_text}
                </div>"""
            body += "</div>"

        body += "<br><p>Thanks.</p><p>Best Regards,<br>Chanchol Roy<br>QED Department</p></body></html>"
        return body

    def run(self, source_files_list: List[Path]) -> Tuple[int, int]:
        """Executes the email automation, drafts emails, and copies review files."""
        logging.info("=" * 50 + "\nTASK 2: Starting Email Automation...\n" + "=" * 50)
        if not source_files_list:
            logging.warning("No files found to email.")
            return 0, 0

        all_reports = [self._get_report_data(f) for f in source_files_list if f]
        all_reports = [r for r in all_reports if r]

        logging.info(f"Found a total of {len(all_reports)} reports. Starting advanced filtering...")

        reports_to_send = []
        reports_to_review = []
        self.review_folder.mkdir(exist_ok=True)

        for report in all_reports:
            if self._classify_report(report) == 'SEND':
                reports_to_send.append(report)
            else:
                reports_to_review.append(report)

        reviewed_count = 0
        for report in reports_to_review:
            try:
                shutil.copy2(report['file_path'], self.review_folder / report['file_path'].name)
                reviewed_count += 1
            except Exception as e:
                logging.error(f"Failed to COPY '{report['file_path'].name}' to review folder: {e}")
        if reviewed_count > 0:
            logging.info(f"{reviewed_count} standard PASS report(s) were COPIED to Manual Review.")

        if not reports_to_send:
            logging.info("Filtering complete. No critical reports found to be sent via email.")
            return 0, reviewed_count

        logging.info(f"Filtering complete. {len(reports_to_send)} email drafts will be created.")
        grouped_reports = defaultdict(list)
        for report in reports_to_send:
            grouped_reports[(report.get('buyer', 'N/A'), report.get('supplier', 'N/A'))].append(report)

        drafts_created = 0
        try:
            outlook = win32.Dispatch('outlook.application')
            for (buyer, supplier), reports in grouped_reports.items():
                has_fail_report = any(
                    'fail' in r.get('result', '').lower() or 'rejected' in r.get('result', '').lower() for r in reports)
                recipient = self.primary_recipient if has_fail_report else self.secondary_recipient
                consignments = sorted(list({r.get('consignment', '') for r in reports if r.get('consignment', '')}))
                subject = f"{buyer} # {', '.join(consignments)} Rolls consignment Fabric Inspection Status"
                body = self._generate_email_body(buyer, supplier, reports)
                attachments = [str(r['file_path']) for r in reports]

                mail = outlook.CreateItem(0)
                mail.To = recipient
                mail.Subject = subject
                mail.HTMLBody = body
                for attachment in attachments: mail.Attachments.Add(attachment)
                mail.Save()
                drafts_created += 1
                logging.info(f"Successfully saved email draft for '{recipient}' with {len(attachments)} attachment(s).")
        except Exception as e:
            logging.error(f"Failed to save email draft using Outlook: {e}")
            logging.error("Please ensure the Outlook application is running.")

        return drafts_created, reviewed_count


# =============================================================================
#   TASK 3: File Organizer
# =============================================================================

class FileOrganizer:
    """Handles file organization and cleanup of empty directories."""

    def __init__(self, config: ConfigLoader):
        self.config = config
        self.source_dir = Path(config.get('paths.pending_reports'))
        self.output_dir = Path(config.get('paths.ongoing_work'))
        self.error_dir = self.source_dir.parent / "Error Reports"
        self.cell_map = config.get('cell_map_organization')

    def _clean_name(self, name: Any) -> str:
        """Cleans a string to be used as a valid file/folder name."""
        name_str = str(name or '').strip()
        return re.sub(r'[\\/*?:"<>|]', '', name_str)

    def _format_date(self, date_value: Any) -> str:
        """Formats a date value into a specific string format."""
        if isinstance(date_value, datetime):
            return date_value.strftime("(%d-%m-%y)")
        try:
            return pd.to_datetime(date_value).strftime("(%d-%m-%y)")
        except (ValueError, TypeError):
            return f"({self._clean_name(date_value)})"

    def _cleanup_empty_dirs(self, path: Path):
        """Recursively deletes empty subdirectories."""
        for dirpath, _, _ in os.walk(path, topdown=False):
            if Path(dirpath) == path: continue
            try:
                os.rmdir(dirpath)
                logging.info(f"   -> Cleaned up empty folder: {dirpath}")
            except OSError:
                pass  # Directory is not empty

    def run(self, source_files_list: List[Path]) -> int:
        """Executes the file organization process."""
        logging.info("=" * 50 + "\nTASK 3: Starting File Organization...\n" + "=" * 50)
        self.error_dir.mkdir(exist_ok=True)
        self.output_dir.mkdir(exist_ok=True)

        if not source_files_list:
            logging.warning("No files to organize.")
            return 0

        organized_count = 0
        for file_path in source_files_list:
            if not file_path.exists():
                continue

            wb = None
            try:
                logging.info(f"\n[*] Organizing file: {file_path.name}")
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                sheet = wb[self.cell_map['sheet_name']]

                data = {key: self._clean_name(sheet[cell].value) for key, cell in self.cell_map.items() if
                        key != 'sheet_name'}
                data['date'] = self._format_date(sheet[self.cell_map['date']].value)

                try:
                    consignment_val = sheet[self.cell_map['consignment']].value
                    data['consignment'] = self._clean_name(int(safe_float(consignment_val)))
                except (ValueError, TypeError):
                    data['consignment'] = self._clean_name(consignment_val)

                try:
                    rolls_val = sheet[self.cell_map['rolls']].value
                    data['rolls'] = self._clean_name(int(safe_float(rolls_val)))
                except (ValueError, TypeError):
                    data['rolls'] = self._clean_name(rolls_val)

                wb.close()
                wb = None

                if not all([data['buyer'], data['supplier'], data['consignment']]):
                    raise ValueError("Missing required info: Buyer, Supplier, or Consignment.")

                folder_name = f"CON-{data['consignment']} {data['date']}"
                destination_folder = self.output_dir / data['buyer'] / folder_name
                destination_folder.mkdir(parents=True, exist_ok=True)

                new_filename = f"{data['style']}, COLOR-{data['color']}, Roll-{data['rolls']}, {data['fabric_code']}{file_path.suffix}"
                final_path = destination_folder / new_filename

                shutil.move(str(file_path), str(final_path))
                logging.info(f"   [✓] Successfully MOVED to '{final_path.relative_to(self.output_dir.parent)}'.")
                organized_count += 1
            except Exception as e:
                logging.error(f"   [X] Error organizing '{file_path.name}': {e}")
                if wb: wb.close()
                try:
                    shutil.move(str(file_path), str(self.error_dir / file_path.name))
                except Exception as move_error:
                    logging.error(f"Could not even move to Error folder: {move_error}")

        logging.info("\nCleaning up empty source folders...")
        self._cleanup_empty_dirs(self.source_dir)
        logging.info("File organization completed.")
        return organized_count


# ==========================================================
#                         Main Driver
# ==========================================================

class AutomationSystem:
    """Manages the entire automation workflow."""

    def __init__(self):
        setup_logging()
        self.config = ConfigLoader(file_path="master.json")
        self.data_entry_handler = DataEntryHandler(self.config)
        self.emailer = Emailer(self.config)
        self.file_organizer = FileOrganizer(self.config)

        self.pending_path = Path(self.config.get('paths.pending_reports'))
        self.ongoing_path = Path(self.config.get('paths.ongoing_work'))

    def _display_menu(self):
        """Displays the main menu to the user."""
        print("\n" + "=" * 60)
        print("    DEBONAIR GROUP - Quality Automation System (v17.4)")
        print("    Created by Chanchol Roy")
        print("=" * 60)
        print("Which task would you like to perform? Please enter a number:")
        print("  1. Run Full Process (Data Entry -> Email -> Organize)")
        print("  2. Only Data Entry")
        print("  3. Only Create Email Drafts")
        print("  4. Only Organize Files")
        print("  0. Exit Program")
        print("=" * 60)

    def _get_files_for_task(self, task_name: str) -> Optional[List[Path]]:
        """Asks the user to choose a folder and returns the files from it."""
        print(f"\nFor the '{task_name}' task, which folder do you want to use?")
        print("  1. Pending Reports")
        print("  2. Ongoing Work")
        folder_choice = input("Your choice (1 or 2): ").strip()

        if folder_choice == '1':
            target_path = self.pending_path
        elif folder_choice == '2':
            target_path = self.ongoing_path
        else:
            print("❌ Invalid choice. Please enter 1 or 2.")
            return None

        files = get_excel_files(target_path)
        if not files:
            logging.error(f"No Excel files found in the '{target_path.name}' folder.")
            return None
        return files

    def run(self):
        """Runs the main loop of the automation system."""
        while True:
            self._display_menu()
            choice = input("Your choice: ").strip()

            if choice == '1':
                files_in_pending = get_excel_files(self.pending_path)
                if not files_in_pending:
                    logging.error(
                        f"Cannot run the full process because the '{self.pending_path.name}' folder is empty.")
                else:
                    initial_files = list(files_in_pending)

                    rows_entered = self.data_entry_handler.run(initial_files)
                    drafts_created, reviewed_count = self.emailer.run(initial_files)
                    organized_count = self.file_organizer.run(initial_files)

                    print("\n" + "-" * 25 + " ACTION SUMMARY " + "-" * 25)
                    print(f"  - Initial Files Found: {len(initial_files)}")
                    print(f"  - Rows Entered: {rows_entered}")
                    print(f"  - Email Drafts Created: {drafts_created}")
                    print(f"  - Files Copied to Review: {reviewed_count}")
                    print(f"  - Files Organized: {organized_count}")
                    print("-" * 68)

            elif choice == '2':
                files = self._get_files_for_task("Data Entry")
                if files:
                    rows_entered = self.data_entry_handler.run(files)
                    print(f"\nSUMMARY: {rows_entered} row(s) were entered into the main workbook.")

            elif choice == '3':
                files = self._get_files_for_task("Email Drafts")
                if files:
                    drafts_created, reviewed_count = self.emailer.run(files)
                    print(f"\nSUMMARY: {drafts_created} email draft(s) were created.")
                    print(f"         {reviewed_count} file(s) were copied to the Manual Review folder.")

            elif choice == '4':
                files = self._get_files_for_task("File Organization")
                if files:
                    organized_count = self.file_organizer.run(files)
                    print(f"\nSUMMARY: {organized_count} file(s) were organized.")

            elif choice == '0':
                print("Exiting program. Goodbye!")
                break
            else:
                print("❌ Invalid input. Please enter a number between 0 and 4.")

            input("\nPress Enter to return to the main menu...")


if __name__ == "__main__":
    try:
        system = AutomationSystem()
        system.run()
    except Exception as e:
        logging.critical(f"A critical error occurred while running the program: {e}", exc_info=True)
        input("\nPress Enter to exit the program...")

